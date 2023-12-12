"""
catxl.py - Excel Workbook Merger

This script merges multiple Excel workbooks from a specified directory into
a single workbook.

Usage:
    python catxl.py <directory>

Arguments:
    directory (str): Path to the directory containing Excel workbooks to be
                     merged.

Example:
    python catxl.py /path/to/excel_files

The script creates a new workbook and copies the contents of each sheet from
each Excel workbook in the specified directory. The new sheets are named using
the original workbook name, and the final merged workbook is saved in the same
directory with the name 'merged_workbook.xlsx'.
"""

from argparse import ArgumentParser
from pathlib import Path

from openpyxl import Workbook, load_workbook


def main() -> None:
    """
    Main function to parse command line arguments and initiate the merging
    process.

    Returns:
        None
    """
    parser = ArgumentParser(
        description="Merge Excel sheets from a directory into a single"
                    " workbook."
    )
    parser.add_argument(
        "directory",
        type=str,
        help="Path to the directory containing Excel workbooks."
    )
    args = parser.parse_args()
    directory = Path(args.directory)

    # Create a new workbook
    merged_workbook = Workbook()

    # Remove the default sheet used for saving the workbook
    merged_workbook.remove(merged_workbook[merged_workbook.sheetnames[0]])

    # Process workbooks in the specified directory
    for workbook_path in directory.glob("*.xlsx"):
        # Load the workbook
        source_workbook = load_workbook(workbook_path)

        # Get the sheet name
        sheet_name = source_workbook.sheetnames[0]

        # Copy the sheet to the new workbook
        source_sheet = source_workbook[sheet_name]
        new_sheet = merged_workbook.create_sheet(title=sheet_name)
        for row in source_sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        # Use the original workbook name in the sheet name
        new_sheet.title = workbook_path.stem

    # Save the new workbook
    merged_workbook.save(directory.name + ".xlsx")


if __name__ == "__main__":
    main()
