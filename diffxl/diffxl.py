"""
diffxl.py - Excel Workbook Difference Extractor

This script finds differences between corresponding cells in each worksheet
of two Excel workbooks and outputs them to a new workbook.
The new workbook's file name incorporates the name of the original workbook.
For example, comparing workbook1.xlsx and workbook2.xlsx produces
'diff_workbook1_workbook2.xlsx'.
The new workbook contains one worksheet with a table representing the
worksheet names, cell addresses, values from the first workbook, and values
from the second workbook where differences were found.
Worksheets that do not exist in one workbook are ignored. Additionally,
hidden cells, blank cells, error cells, and formula cells are evaluated as
empty strings.

Usage:
    python diffxl.py [-h] workbook1.xlsx workbook2.xlsx

Arguments:
    path1 (str): Path to the first Excel workbook.
    path2 (str): Path to the second Excel workbook or a director containing
                 Excel workbooks.
    -h, --help (flag): Show this help message and exit.
"""

from argparse import ArgumentParser
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.cell.cell import (TYPE_ERROR, TYPE_FORMULA,
                                TYPE_FORMULA_CACHE_STRING, TYPE_NULL)
from openpyxl.worksheet.worksheet import Worksheet


def is_worksheet_hidden(worksheet: Worksheet) -> bool:
    """Check if a worksheet is hidden."""
    return worksheet.sheet_state in ["hidden", "veryHidden"]


def is_cell_hidden(worksheet: Worksheet, cell: Cell) -> bool:
    """Check if a cell is hidden."""
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Check if any cell in the merged range is hidden.
            for merged_cell in merged_range:
                column = merged_cell.column_letter
                row = merged_cell.row
                if (
                    worksheet.column_dimensions[column].hidden
                    or worksheet.row_dimensions[row].hidden
                ):
                    return True
            return False
    return False


def main() -> None:
    """
    Main function to extract differences from two Excel workbooks and
    save to a new workbook using command line arguments.
    """
    # Read commandline arguments.
    parser = ArgumentParser(
        description="Extract differences from two Excel workbooks and"
        " save to a new workbook."
    )
    parser.add_argument("path1", type=str, help="Path to the first workbook.")
    parser.add_argument(
        "path2",
        type=str,
        help="Path to the second workbook or a directory containing workbooks.",
    )
    args = vars(parser.parse_args())
    path1 = Path(args["path1"])
    path2 = Path(args["path2"])

    # Display first workbook name in the console.
    print(f"Opening '{path1.name}'...")

    # Load the first Excel workbooks
    workbook1 = load_workbook(path1)

    # Extract differences and write to file.
    # If the second argument is a directory, process all Excel workbooks
    # in that directory, otherwise process a single Excel workbook.
    difference_found = False
    for target in list(path2.glob("*.xlsx") if path2.is_dir() else [path2]):
        # Initialize list of differences.
        differences = []

        # Display second workbook name in the console.
        print(f"Opening '{target.name}'...")

        # Load the second Excel workbooks
        workbook2 = load_workbook(target)

        # Iterate through each worksheet
        for sheetname in workbook1.sheetnames:
            # Exclude non-existent worksheets.
            if sheetname not in workbook2.sheetnames:
                continue

            worksheet1: Worksheet = workbook1[sheetname]
            worksheet2: Worksheet = workbook2[sheetname]

            # Exclude hidden worksheets.
            if is_worksheet_hidden(worksheet1):
                continue
            if is_worksheet_hidden(worksheet2):
                continue

            # Display worksheet and workbook names in the console.
            print(f"Comparing '{sheetname}' in '{target.name}'...")

            # Get the max row and column for the current worksheet.
            max_row = max(worksheet1.max_row, worksheet2.max_row)
            max_col = max(worksheet1.max_column, worksheet2.max_column)

            # Iterate through each cell in the worksheet.
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell1 = worksheet1.cell(row, col)
                    cell2 = worksheet2.cell(row, col)
                    value1 = str(cell1.value)
                    value2 = str(cell2.value)

                    # Treat values ​​in hidden cells, blank cells, error cells,
                    # and formula cells as empty strings.
                    if is_cell_hidden(worksheet1, cell1):
                        value1 = ""
                    if cell1.data_type in [
                        TYPE_NULL,
                        TYPE_ERROR,
                        TYPE_FORMULA,
                        TYPE_FORMULA_CACHE_STRING,
                    ]:
                        value1 = ""
                    if is_cell_hidden(worksheet2, cell2):
                        value2 = ""
                    if cell2.data_type in [
                        TYPE_NULL,
                        TYPE_ERROR,
                        TYPE_FORMULA,
                        TYPE_FORMULA_CACHE_STRING,
                    ]:
                        value2 = ""

                    # Compare cell values.
                    if value1 != value2:
                        differences.append({
                            "worksheet_name": worksheet1.title,
                            "cell_address": cell1.coordinate,
                            "value1": value1,
                            "value2": value2,
                        })

        # Close the second workbook.
        workbook2.close()

        if differences:
            # Initialize differences workbook.
            print("Writing...")
            diff_workbook = Workbook()
            diff_worksheet: Worksheet = diff_workbook[diff_workbook.sheetnames[0]]
            diff_worksheet.title = "Differences"

            # Add headers for worksheet name, cell address, and values columns.
            diff_worksheet["A1"] = "Sheet"
            diff_worksheet["B1"] = "Cell"
            diff_worksheet["C1"] = f"{path1.stem} Value"
            diff_worksheet["D1"] = f"{target.stem} Value"

            # Iterate through differences and write to the worksheet.
            for row, difference in enumerate(differences, 2):
                diff_worksheet.cell(row, 1, difference["worksheet_name"])
                diff_worksheet.cell(row, 2, difference["cell_address"])
                diff_worksheet.cell(row, 3, difference["value1"])
                diff_worksheet.cell(row, 4, difference["value2"])

            # Save the differences workbook.
            diff_workbook_name = f"diff_{path1.stem}_{target.stem}.xlsx"
            diff_workbook.save(diff_workbook_name)
            print(f"Differences saved to {diff_workbook_name}")

        else:
            print("No differences found.")

    # Close the first workbook.
    workbook1.close()

    print("Done.")


if __name__ == "__main__":
    main()
